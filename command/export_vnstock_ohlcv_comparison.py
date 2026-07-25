"""Xuất Excel so sánh OHLCV từ KBS và VCI cho các phiên gần nhất.

Mặc định:
    python command/export_vnstock_ohlcv_comparison.py

Tùy chỉnh:
    python command/export_vnstock_ohlcv_comparison.py --symbols VNINDEX SAB DRI --sessions 10
"""

from __future__ import annotations

import argparse
import sys
import time
from dataclasses import dataclass
from datetime import date
from importlib.metadata import PackageNotFoundError, version
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT_DIR = Path(__file__).resolve().parent.parent
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

SOURCES = ("KBS", "VCI")
FIELDS = ("Open", "High", "Low", "Close", "Volume")
REQUIRED_COLUMNS = ("time", "open", "high", "low", "close", "volume")
DEFAULT_SYMBOLS = ("VNINDEX", "SAB", "DRI")


@dataclass
class FetchResult:
    symbol: str
    source: str
    status: str
    api_rows: int
    elapsed_seconds: float
    first_date: str | None
    last_date: str | None
    error: str | None
    data: pd.DataFrame


def _configure_utf8_console() -> None:
    for stream in (sys.stdout, sys.stderr):
        reconfigure = getattr(stream, "reconfigure", None)
        if callable(reconfigure):
            reconfigure(encoding="utf-8")


def _package_version(package: str) -> str:
    try:
        return version(package)
    except PackageNotFoundError:
        return "unknown"


def normalize_daily_ohlcv(
    frame: pd.DataFrame, symbol: str, source: str
) -> pd.DataFrame:
    """Chuẩn hóa response vnstock và giữ cả timestamp gốc để đối chiếu."""
    if frame is None or frame.empty:
        raise ValueError("API không trả về dữ liệu")
    missing = [column for column in REQUIRED_COLUMNS if column not in frame.columns]
    if missing:
        raise ValueError(f"Thiếu cột bắt buộc: {', '.join(missing)}")

    output = pd.DataFrame(
        {
            "Ma": symbol,
            "Nguon": source,
            "Thoi_gian_goc": pd.to_datetime(frame["time"], errors="coerce"),
            "Open": pd.to_numeric(frame["open"], errors="coerce"),
            "High": pd.to_numeric(frame["high"], errors="coerce"),
            "Low": pd.to_numeric(frame["low"], errors="coerce"),
            "Close": pd.to_numeric(frame["close"], errors="coerce"),
            "Volume": pd.to_numeric(frame["volume"], errors="coerce"),
        }
    )
    if output["Thoi_gian_goc"].isna().any():
        raise ValueError("Có timestamp không đọc được")
    if output.loc[:, list(FIELDS[:4])].isna().any(axis=None):
        raise ValueError("Có giá trị OHLC rỗng")

    output["Ngay"] = output["Thoi_gian_goc"].dt.normalize()
    output = output.sort_values("Thoi_gian_goc").drop_duplicates("Ngay", keep="last")
    return output.loc[:, ["Ma", "Nguon", "Thoi_gian_goc", "Ngay", *FIELDS]].reset_index(
        drop=True
    )


def fetch_daily_ohlcv(*, symbol: str, source: str, start: str, end: str) -> FetchResult:
    started = time.perf_counter()
    try:
        from vnstock import Quote

        raw = Quote(symbol=symbol, source=source).history(
            start=start,
            end=end,
            interval="1D",
        )
        api_rows = 0 if raw is None else len(raw)
        normalized = normalize_daily_ohlcv(raw, symbol, source)
        elapsed = round(time.perf_counter() - started, 3)
        return FetchResult(
            symbol=symbol,
            source=source,
            status="PASS",
            api_rows=api_rows,
            elapsed_seconds=elapsed,
            first_date=normalized["Ngay"].min().date().isoformat(),
            last_date=normalized["Ngay"].max().date().isoformat(),
            error=None,
            data=normalized,
        )
    except Exception as exc:
        return FetchResult(
            symbol=symbol,
            source=source,
            status="FAIL",
            api_rows=0,
            elapsed_seconds=round(time.perf_counter() - started, 3),
            first_date=None,
            last_date=None,
            error=f"{type(exc).__name__}: {exc}",
            data=pd.DataFrame(
                columns=["Ma", "Nguon", "Thoi_gian_goc", "Ngay", *FIELDS]
            ),
        )


def build_comparison(
    results: list[FetchResult], sessions: int
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Tạo bảng đặt KBS/VCI cạnh nhau trên hợp các ngày mới nhất."""
    comparison_rows: list[dict[str, object]] = []
    summary_rows: list[dict[str, object]] = []
    selected_by_source: dict[str, list[pd.DataFrame]] = {
        source: [] for source in SOURCES
    }
    symbols = list(dict.fromkeys(item.symbol for item in results))

    for symbol in symbols:
        by_source = {
            item.source: item
            for item in results
            if item.symbol == symbol and item.source in SOURCES
        }
        dates = pd.DatetimeIndex([])
        for item in by_source.values():
            dates = dates.union(pd.DatetimeIndex(item.data["Ngay"]))
        recent_dates = dates.sort_values()[-sessions:]

        indexed: dict[str, pd.DataFrame] = {}
        for source in SOURCES:
            item = by_source.get(source)
            if item is None or item.data.empty:
                indexed[source] = pd.DataFrame(index=recent_dates, columns=FIELDS)
                continue
            chosen = item.data[item.data["Ngay"].isin(recent_dates)].copy()
            selected_by_source[source].append(chosen)
            indexed[source] = chosen.set_index("Ngay")

        mismatch_counts = {field: 0 for field in FIELDS}
        fully_matching = 0
        missing_kbs = 0
        missing_vci = 0
        for session_date in recent_dates:
            row: dict[str, object] = {"Ma": symbol, "Ngay": session_date}
            present: dict[str, bool] = {}
            for source in SOURCES:
                frame = indexed[source]
                present[source] = session_date in frame.index
                row[f"Du_{source}"] = present[source]
                for field in FIELDS:
                    row[f"{source}_{field}"] = (
                        frame.at[session_date, field] if present[source] else np.nan
                    )

            if not present["KBS"]:
                missing_kbs += 1
            if not present["VCI"]:
                missing_vci += 1

            differing_fields: list[str] = []
            for field in FIELDS:
                kbs_value = row[f"KBS_{field}"]
                vci_value = row[f"VCI_{field}"]
                if pd.notna(kbs_value) and pd.notna(vci_value):
                    difference = float(vci_value) - float(kbs_value)
                    difference_pct = (
                        difference / float(kbs_value)
                        if float(kbs_value) != 0
                        else np.nan
                    )
                    matches = bool(
                        np.isclose(
                            float(vci_value),
                            float(kbs_value),
                            rtol=0.0,
                            atol=1e-9,
                        )
                    )
                else:
                    difference = np.nan
                    difference_pct = np.nan
                    matches = False
                row[f"Chenh_{field}"] = difference
                row[f"Chenh_pct_{field}"] = difference_pct
                row[f"Khop_{field}"] = matches
                if not matches:
                    mismatch_counts[field] += 1
                    differing_fields.append(field)

            row["Khop_tat_ca"] = not differing_fields
            row["Cot_sai_khac"] = ", ".join(differing_fields)
            if not differing_fields:
                fully_matching += 1
            comparison_rows.append(row)

        kbs_result = by_source.get("KBS")
        vci_result = by_source.get("VCI")
        summary_rows.append(
            {
                "Ma": symbol,
                "So_phien_xuat": len(recent_dates),
                "Ngay_dau": recent_dates.min() if len(recent_dates) else pd.NaT,
                "Ngay_cuoi": recent_dates.max() if len(recent_dates) else pd.NaT,
                "KBS_Trang_thai": kbs_result.status if kbs_result else "MISSING",
                "KBS_So_dong_API": kbs_result.api_rows if kbs_result else 0,
                "KBS_Loi": kbs_result.error if kbs_result else "Không có kết quả",
                "VCI_Trang_thai": vci_result.status if vci_result else "MISSING",
                "VCI_So_dong_API": vci_result.api_rows if vci_result else 0,
                "VCI_Loi": vci_result.error if vci_result else "Không có kết quả",
                "Thieu_KBS": missing_kbs,
                "Thieu_VCI": missing_vci,
                "Phien_khop_tat_ca": fully_matching,
                **{f"Sai_{field}": count for field, count in mismatch_counts.items()},
            }
        )

    comparison = pd.DataFrame(comparison_rows)
    summary = pd.DataFrame(summary_rows)
    raw_frames: dict[str, pd.DataFrame] = {}
    for source in SOURCES:
        raw_frames[source] = (
            pd.concat(selected_by_source[source], ignore_index=True)
            if selected_by_source[source]
            else pd.DataFrame(columns=["Ma", "Nguon", "Thoi_gian_goc", "Ngay", *FIELDS])
        )
    return summary, comparison, raw_frames["KBS"], raw_frames["VCI"]


def _style_worksheet(worksheet) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    worksheet.freeze_panes = "A2"
    if worksheet.max_row >= 1 and worksheet.max_column >= 1:
        worksheet.auto_filter.ref = worksheet.dimensions

    for column_index, cells in enumerate(worksheet.columns, start=1):
        max_length = max(len(str(cell.value or "")) for cell in cells)
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(
            max(max_length + 2, 11), 34
        )

    headers = {cell.value: cell.column for cell in worksheet[1]}
    for header, column_index in headers.items():
        header_text = str(header or "")
        for row_index in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row=row_index, column=column_index)
            if (
                header_text == "Ngay"
                or header_text.endswith("_dau")
                or header_text.endswith("_cuoi")
            ):
                cell.number_format = "dd/mm/yyyy"
            elif "Volume" in header_text or "So_dong" in header_text:
                cell.number_format = "#,##0"
            elif "pct" in header_text:
                cell.number_format = "0.0000%"
            elif any(field in header_text for field in FIELDS[:4]):
                cell.number_format = "#,##0.00"


def _highlight_report(workbook) -> None:
    """Highlight every KBS/VCI value pair that does not match."""
    red_fill = PatternFill("solid", fgColor="F4CCCC")
    yellow_fill = PatternFill("solid", fgColor="FFF2CC")
    orange_fill = PatternFill("solid", fgColor="FCE5CD")
    green_fill = PatternFill("solid", fgColor="D9EAD3")

    compare_sheet = workbook["So_sanh"]
    headers = {cell.value: cell.column for cell in compare_sheet[1]}
    for row_index in range(2, compare_sheet.max_row + 1):
        all_match_cell = compare_sheet.cell(
            row=row_index, column=headers["Khop_tat_ca"]
        )
        all_match_cell.fill = green_fill if all_match_cell.value is True else red_fill

        for header in ("Du_KBS", "Du_VCI"):
            present_cell = compare_sheet.cell(row=row_index, column=headers[header])
            present_cell.fill = green_fill if present_cell.value is True else red_fill

        differing_fields: list[str] = []
        for field in FIELDS:
            match_cell = compare_sheet.cell(
                row=row_index, column=headers[f"Khop_{field}"]
            )
            if match_cell.value is True:
                match_cell.fill = green_fill
                continue

            differing_fields.append(field)
            match_cell.fill = red_fill
            compare_sheet.cell(
                row=row_index, column=headers[f"KBS_{field}"]
            ).fill = yellow_fill
            compare_sheet.cell(
                row=row_index, column=headers[f"VCI_{field}"]
            ).fill = red_fill
            compare_sheet.cell(
                row=row_index, column=headers[f"Chenh_{field}"]
            ).fill = orange_fill
            compare_sheet.cell(
                row=row_index, column=headers[f"Chenh_pct_{field}"]
            ).fill = orange_fill

        if differing_fields:
            compare_sheet.cell(row=row_index, column=headers["Ma"]).fill = red_fill
            compare_sheet.cell(row=row_index, column=headers["Ngay"]).fill = red_fill
            compare_sheet.cell(
                row=row_index, column=headers["Cot_sai_khac"]
            ).fill = red_fill

    summary_sheet = workbook["Tong_quan"]
    summary_headers = {cell.value: cell.column for cell in summary_sheet[1]}
    count_headers = ["Thieu_KBS", "Thieu_VCI", *[f"Sai_{field}" for field in FIELDS]]
    for row_index in range(2, summary_sheet.max_row + 1):
        has_difference = False
        for header in count_headers:
            cell = summary_sheet.cell(row=row_index, column=summary_headers[header])
            if isinstance(cell.value, (int, float)) and cell.value > 0:
                cell.fill = red_fill
                has_difference = True
            else:
                cell.fill = green_fill

        symbol_cell = summary_sheet.cell(row=row_index, column=summary_headers["Ma"])
        matched_cell = summary_sheet.cell(
            row=row_index, column=summary_headers["Phien_khop_tat_ca"]
        )
        symbol_cell.fill = red_fill if has_difference else green_fill
        matched_cell.fill = red_fill if has_difference else green_fill

    guide_sheet = workbook["Huong_dan"]
    guide_headers = {cell.value: cell.column for cell in guide_sheet[1]}
    legend_fills = {
        "Mau_vang": yellow_fill,
        "Mau_do": red_fill,
        "Mau_cam": orange_fill,
        "Mau_xanh": green_fill,
    }
    for row_index in range(2, guide_sheet.max_row + 1):
        key = guide_sheet.cell(row=row_index, column=guide_headers["Muc"]).value
        if key in legend_fills:
            guide_sheet.cell(
                row=row_index, column=guide_headers["Gia_tri"]
            ).fill = legend_fills[key]


def write_workbook(
    output_path: Path,
    *,
    summary: pd.DataFrame,
    comparison: pd.DataFrame,
    kbs_raw: pd.DataFrame,
    vci_raw: pd.DataFrame,
    metadata: dict[str, object],
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    guide = pd.DataFrame(
        {
            "Muc": list(metadata.keys()),
            "Gia_tri": [str(value) for value in metadata.values()],
        }
    )
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        guide.to_excel(writer, sheet_name="Huong_dan", index=False)
        summary.to_excel(writer, sheet_name="Tong_quan", index=False)
        comparison.to_excel(writer, sheet_name="So_sanh", index=False)
        kbs_raw.to_excel(writer, sheet_name="KBS_raw", index=False)
        vci_raw.to_excel(writer, sheet_name="VCI_raw", index=False)

        for worksheet in writer.book.worksheets:
            _style_worksheet(worksheet)
        _highlight_report(writer.book)


def export_comparison(
    *,
    symbols: list[str],
    sessions: int,
    end: str,
    lookback_days: int,
    pause_seconds: float,
    output_path: Path,
) -> tuple[Path, list[FetchResult]]:
    start = (pd.Timestamp(end) - pd.Timedelta(days=lookback_days)).date().isoformat()
    results: list[FetchResult] = []
    for symbol in symbols:
        for source in SOURCES:
            print(f"[FETCH] {symbol} / {source} / {start}..{end}")
            result = fetch_daily_ohlcv(
                symbol=symbol,
                source=source,
                start=start,
                end=end,
            )
            results.append(result)
            print(
                f"  [{result.status}] rows={result.api_rows}, "
                f"range={result.first_date}..{result.last_date}, "
                f"elapsed={result.elapsed_seconds:.3f}s"
            )
            if result.error:
                print(f"  ERROR: {result.error}")
            if pause_seconds > 0:
                time.sleep(pause_seconds)

    summary, comparison, kbs_raw, vci_raw = build_comparison(results, sessions)
    metadata = {
        "Thoi_gian_xuat": pd.Timestamp.now(tz="Asia/Ho_Chi_Minh").isoformat(),
        "Phien_ban_vnstock": _package_version("vnstock"),
        "Ma_chung_khoan": ", ".join(symbols),
        "So_phien": sessions,
        "Khoang_goi_API": f"{start} đến {end}",
        "Huong_chenh_lech": "Chenh_* = VCI - KBS",
        "Mau_vang": "Gia tri KBS cua truong dang sai khac",
        "Mau_do": "Gia tri VCI, co sai hoac ket qua khong khop",
        "Mau_cam": "Gia tri chenh lech VCI - KBS",
        "Mau_xanh": "Du lieu hoac ket qua khop",
        "Quy_tac_chon_phien": "Hợp các ngày mới nhất của KBS và VCI theo từng mã",
        "Luu_y": "Timestamp daily được chuẩn hóa về ngày; timestamp gốc nằm trong sheet raw",
    }
    write_workbook(
        output_path,
        summary=summary,
        comparison=comparison,
        kbs_raw=kbs_raw,
        vci_raw=vci_raw,
        metadata=metadata,
    )
    return output_path, results


def _parse_args() -> argparse.Namespace:
    today = date.today()
    parser = argparse.ArgumentParser(
        description="Xuất Excel so sánh OHLCV daily giữa KBS và VCI."
    )
    parser.add_argument("--symbols", nargs="+", default=list(DEFAULT_SYMBOLS))
    parser.add_argument("--sessions", type=int, default=10)
    parser.add_argument("--end", default=today.isoformat())
    parser.add_argument("--lookback-days", type=int, default=45)
    parser.add_argument("--pause", type=float, default=0.5)
    parser.add_argument("--output", type=Path)
    return parser.parse_args()


def main() -> int:
    _configure_utf8_console()
    args = _parse_args()
    symbols = [str(symbol).strip().upper() for symbol in args.symbols]
    if args.sessions < 1:
        raise SystemExit("--sessions phải lớn hơn 0")
    output_path = args.output or (
        ROOT_DIR
        / "reports"
        / f"KBS_VCI_OHLCV_{args.sessions}_phien_{pd.Timestamp(args.end):%Y%m%d}.xlsx"
    )
    path, results = export_comparison(
        symbols=symbols,
        sessions=args.sessions,
        end=args.end,
        lookback_days=max(args.lookback_days, args.sessions * 2),
        pause_seconds=max(args.pause, 0.0),
        output_path=output_path,
    )
    print(f"[DONE] Excel: {path.resolve()}")
    return 1 if any(result.status == "FAIL" for result in results) else 0


if __name__ == "__main__":
    raise SystemExit(main())
