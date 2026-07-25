from __future__ import annotations

import pandas as pd
from openpyxl import load_workbook

from command.export_vnstock_ohlcv_comparison import (
    FetchResult,
    build_comparison,
    normalize_daily_ohlcv,
    write_workbook,
)


def _api_frame(times, opens, volumes) -> pd.DataFrame:
    return pd.DataFrame(
        {
            "time": times,
            "open": opens,
            "high": [value + 1 for value in opens],
            "low": [value - 1 for value in opens],
            "close": [value + 0.5 for value in opens],
            "volume": volumes,
        }
    )


def _result(symbol: str, source: str, frame: pd.DataFrame) -> FetchResult:
    normalized = normalize_daily_ohlcv(frame, symbol, source)
    return FetchResult(
        symbol=symbol,
        source=source,
        status="PASS",
        api_rows=len(frame),
        elapsed_seconds=0.1,
        first_date=normalized["Ngay"].min().date().isoformat(),
        last_date=normalized["Ngay"].max().date().isoformat(),
        error=None,
        data=normalized,
    )


def test_build_comparison_keeps_union_dates_and_calculates_vci_minus_kbs():
    kbs = _api_frame(
        ["2026-07-23 07:00:00", "2026-07-24 07:00:00"],
        [100.0, 101.0],
        [1_000, 2_000],
    )
    vci = _api_frame(
        ["2026-07-22", "2026-07-23", "2026-07-24"],
        [99.0, 100.0, 102.0],
        [900, 1_000, 2_100],
    )

    summary, comparison, kbs_raw, vci_raw = build_comparison(
        [_result("SAB", "KBS", kbs), _result("SAB", "VCI", vci)],
        sessions=3,
    )

    assert comparison["Ngay"].dt.strftime("%Y-%m-%d").tolist() == [
        "2026-07-22",
        "2026-07-23",
        "2026-07-24",
    ]
    last = comparison.iloc[-1]
    assert last["Chenh_Open"] == 1.0
    assert last["Chenh_Volume"] == 100.0
    assert last["Khop_tat_ca"] is False or not bool(last["Khop_tat_ca"])
    assert summary.loc[0, "Thieu_KBS"] == 1
    assert len(kbs_raw) == 2
    assert len(vci_raw) == 3


def test_write_workbook_creates_expected_evidence_sheets(tmp_path):
    kbs = _result(
        "DRI",
        "KBS",
        _api_frame(["2026-07-24 07:00:00"], [12.0], [500]),
    )
    vci = _result(
        "DRI",
        "VCI",
        _api_frame(["2026-07-24"], [12.0], [500]),
    )
    summary, comparison, kbs_raw, vci_raw = build_comparison([kbs, vci], sessions=10)
    output = tmp_path / "comparison.xlsx"

    write_workbook(
        output,
        summary=summary,
        comparison=comparison,
        kbs_raw=kbs_raw,
        vci_raw=vci_raw,
        metadata={"So_phien": 10, "Huong_chenh_lech": "VCI - KBS"},
    )

    workbook = load_workbook(output, data_only=True)
    assert workbook.sheetnames == [
        "Huong_dan",
        "Tong_quan",
        "So_sanh",
        "KBS_raw",
        "VCI_raw",
    ]
    assert workbook["So_sanh"].freeze_panes == "A2"
    assert workbook["So_sanh"].max_row == 2


def test_write_workbook_highlights_each_differing_value_pair(tmp_path):
    kbs = _result(
        "VNINDEX",
        "KBS",
        _api_frame(["2026-07-24"], [100.0], [1_000]),
    )
    vci = _result(
        "VNINDEX",
        "VCI",
        _api_frame(["2026-07-24"], [101.0], [1_000]),
    )
    summary, comparison, kbs_raw, vci_raw = build_comparison([kbs, vci], sessions=10)
    output = tmp_path / "highlighted.xlsx"

    write_workbook(
        output,
        summary=summary,
        comparison=comparison,
        kbs_raw=kbs_raw,
        vci_raw=vci_raw,
        metadata={"So_phien": 10},
    )

    workbook = load_workbook(output, data_only=True)
    sheet = workbook["So_sanh"]
    headers = {cell.value: cell.column for cell in sheet[1]}

    assert sheet.cell(2, headers["KBS_Open"]).fill.fgColor.rgb.endswith("FFF2CC")
    assert sheet.cell(2, headers["VCI_Open"]).fill.fgColor.rgb.endswith("F4CCCC")
    assert sheet.cell(2, headers["Chenh_Open"]).fill.fgColor.rgb.endswith("FCE5CD")
    assert sheet.cell(2, headers["Khop_Open"]).fill.fgColor.rgb.endswith("F4CCCC")
    assert sheet.cell(2, headers["Khop_Volume"]).fill.fgColor.rgb.endswith("D9EAD3")

    summary_sheet = workbook["Tong_quan"]
    summary_headers = {cell.value: cell.column for cell in summary_sheet[1]}
    assert summary_sheet.cell(2, summary_headers["Sai_Open"]).fill.fgColor.rgb.endswith(
        "F4CCCC"
    )
